import pandas as pd
import sys
from openpyxl import Workbook

from ITSM_Excel.Excel_Calculate.ChangeRequest.crByCatSubcat import CRByCatSubcat
from ITSM_Excel.Excel_Calculate.ChangeRequest.overallCRSummary import OverallCRSummary
from ITSM_Excel.Excel_Calculate.ChangeRequest.top5CRTypes import Top5CRTypes
from ITSM_Excel.Excel_Calculate.Incident.incByCatSubcat import IncByCatSubcat
from ITSM_Excel.Excel_Calculate.Incident.overallIncidentSummary import OverallIncidentSummary
from ITSM_Excel.Excel_Calculate.Incident.top5IncidentTypes import Top5IncidentTypes
from ITSM_Excel.Excel_Calculate.Request.overallRequestSummary import OverallRequestSummary
from ITSM_Excel.Excel_FetchData.fetchData import FetchData
from ITSM_Util.inputUtil import InputUtil


class WriteFileExcel:
    def __init__(self, input_project_name, input_date_range_file, fetch_data):
        self.fetch_data = fetch_data

        project_name = input_project_name
        date_range = input_date_range_file
        self.output_path_incidents = "Output//GCO - " + project_name + " Weekly ITSM Report - Incident - " + date_range + ".xlsx"
        self.output_path_cr = "Output//GCO - " + project_name + " Weekly ITSM Report - CR - " + date_range + ".xlsx"
        self.output_path_request = "Output//GCO - " + project_name + " Weekly ITSM Report - Request - " + date_range + ".xlsx"


    def write(self):

        try:
            self.write_incidents()
        except Exception as exc:
            print("\n While writing Incidents to Excel : ", exc, "\n")
            wb = Workbook()
            wb.create_sheet("Overall Incident Summary")
            wb.create_sheet("Pie Chart")
            wb.create_sheet("Incidents by Cat and Subcat")
            wb.create_sheet("Top 5 Incident Types")
            wb.save(self.output_path_incidents)

        try:
            self.write_CR()
        except Exception as exc:
            print("\n While writing CRs to Excel : ", exc, "\n")
            wb = Workbook()
            wb.create_sheet("Overall CR Summary")
            wb.create_sheet("CR by Cat and Subcat")
            wb.create_sheet("Top 5 CR Types")
            wb.save(self.output_path_cr)

        try:
            self.write_request()
        except Exception as exc:
            print("\n While writing Requests to Excel : ", exc, "\n")
            wb = Workbook()
            wb.create_sheet("Overall Request Summary")
            wb.save(self.output_path_request)

    def write_incidents(self):

        """Incidents created and resolved by priority"""

        overall_incident_summary = OverallIncidentSummary(self.fetch_data)
        overall_incident_summary_data_list = overall_incident_summary.get_data()
        overall_incident_summary_sheet = pd.DataFrame(overall_incident_summary_data_list[1:],
                                                      columns=overall_incident_summary_data_list[0])

        """Incidents by priority - Pie chart"""
        pie_chart_data = OverallIncidentSummary(self.fetch_data)
        pie_chart_data_list = pie_chart_data.get_piechart_data()
        pie_chart_data_sheet = pd.DataFrame(pie_chart_data_list[1:], columns=pie_chart_data_list[0])

        """Incident by Category Subcategory"""
        inc_by_cat_subcat = IncByCatSubcat(self.fetch_data)
        inc_by_cat_subcat_data_list = inc_by_cat_subcat.get_data()
        inc_by_cat_subcat_sheet = pd.DataFrame(inc_by_cat_subcat_data_list[1:], columns=inc_by_cat_subcat_data_list[0])

        """Top 5 Incident Types"""
        top5_incident_types = Top5IncidentTypes(self.fetch_data)
        top5_incident_types_data_list = top5_incident_types.get_data()
        top5_incident_types_sheet = pd.DataFrame(top5_incident_types_data_list[1:], columns=top5_incident_types_data_list[0])

        """Final Excel Creation"""
        incident_sheets = {'Overall Incident Summary': overall_incident_summary_sheet,
                           'Pie Chart': pie_chart_data_sheet,
                           'Incidents by Cat and Subcat': inc_by_cat_subcat_sheet,
                           'Top 5 Incident Types': top5_incident_types_sheet}

        writer = pd.ExcelWriter(self.output_path_incidents, engine='xlsxwriter')
        for sheet in incident_sheets.keys():
            incident_sheets[sheet].to_excel(writer, sheet_name=sheet, index=False)

        writer.close()


    def write_CR(self):

        """Overall CR Summary"""
        overall_cr_summary = OverallCRSummary(self.fetch_data)
        overall_cr_summary_data_list = overall_cr_summary.get_data()
        overall_cr_summary_sheet = pd.DataFrame(overall_cr_summary_data_list[1:], columns=overall_cr_summary_data_list[0])

        """CR by Category Subcategory"""
        cr_by_cat_subcat = CRByCatSubcat(self.fetch_data)
        cr_by_cat_subcat_data_list = cr_by_cat_subcat.get_data()
        cr_by_cat_subcat_sheet = pd.DataFrame(cr_by_cat_subcat_data_list[1:], columns=cr_by_cat_subcat_data_list[0])

        """Top 5 CR Types"""
        top5_cr_types = Top5CRTypes(self.fetch_data)
        top5_cr_types_data_list = top5_cr_types.get_data()
        top5_cr_types_sheet = pd.DataFrame(top5_cr_types_data_list[1:], columns=top5_cr_types_data_list[0])

        """Final Excel Creation"""
        cr_sheets = {'Overall CR Summary': overall_cr_summary_sheet,
                     'CR by Cat and Subcat': cr_by_cat_subcat_sheet,
                     'Top 5 CR Types': top5_cr_types_sheet}

        writer = pd.ExcelWriter(self.output_path_cr, engine='xlsxwriter')
        for sheet in cr_sheets.keys():
            cr_sheets[sheet].to_excel(writer, sheet_name=sheet, index=False)
        writer.close()


    def write_request(self):

        """Overall Request Summary"""
        overall_request = OverallRequestSummary(self.fetch_data)
        overall_request_data_list = overall_request.get_data()
        overall_request_sheet = pd.DataFrame(overall_request_data_list[1:], columns=overall_request_data_list[0])

        """Final Excel Creation"""
        request_sheets = {'Overall Request Summary': overall_request_sheet}
        writer = pd.ExcelWriter(self.output_path_request, engine='xlsxwriter')

        for sheet in request_sheets.keys():
            request_sheets[sheet].to_excel(writer, sheet_name=sheet, index=False)
        writer.close()

"""
obj = WriteFileExcel()
obj.write_incidents()
obj.write_CR()
obj.write_request()
"""