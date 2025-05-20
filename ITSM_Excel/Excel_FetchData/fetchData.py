import openpyxl
from openpyxl import Workbook
import pandas as pd
import datetime
import sys

from ITSM_Excel.Excel_FetchData.fetchUtils import FetchUtils


class FetchData:

    def __init__(self, input_start_date, input_end_date, input_project_key):

        # store input path files
        self.incident_sheet_path = "ServiceNow Dump\\HP Weekly ITSM Report - Total Incidents Master - Dump.xlsx"
        self.response_resolution_sheet_path = "ServiceNow Dump\\HP Weekly ITSM Report - Response and Resolution Master - Dump.xlsx"
        self.change_request_sheet_path = "ServiceNow Dump\\HP Weekly ITSM Report - Change Request Master - Dump.xlsx"
        self.request_sheet_path = "ServiceNow Dump\\HP Weekly ITSM Report - Requested Item Master - Dump.xlsx"


        # output path for intermediary files
        self.intermediary_incident_sheet_path = "Intermediary\\Intermediary - Incidents.xlsx"
        self.intermediary_response_resolution_sheet_path = "Intermediary\\Intermediary - Response Resolution.xlsx"
        self.intermediary_change_request_sheet_path = "Intermediary\\Intermediary - Change Request.xlsx"
        self.intermediary_request_sheet_path = "Intermediary\\Intermediary - Request Item.xlsx"

        # get start date and end date from user input
        self.start_date = input_start_date
        self.end_date = input_end_date


        # get project key from user input
        self.input_project_key = input_project_key
        self.project_key = self.input_project_key


        self.fetchUtils = FetchUtils()
        self.incident_sheet_data = self.fetch_incident_sheet_data()
        self.response_resolution_sheet_data = self.fetch_response_resolution_sheet_data()
        self.change_request_sheet_data = self.fetch_change_request_sheet_data()
        self.request_sheet_data = self.fetch_request_sheet_data()

        # self.start_date = datetime.datetime(2021, 4, 1)
        # self.end_date = datetime.datetime(2021, 5, 1)

    def get_incident_sheet_data(self):
        return self.incident_sheet_data.copy()

    def get_response_resolution_sheet_data(self):
        return self.response_resolution_sheet_data.copy()

    def get_change_request_sheet_data(self):
        return self.change_request_sheet_data.copy()

    def get_request_sheet_data(self):
        return self.request_sheet_data.copy()

    def fetch_incident_sheet_data(self):

        wb = openpyxl.load_workbook(self.incident_sheet_path, data_only=True)
        ws = wb.active
        L = []
        for row in ws.rows:
            l = []
            for cell in row:
                l.append(cell.value)
            L.append(l)

        df = pd.DataFrame(L[1:], columns=L[0])

        mask1 = (df["Project Key"] == self.project_key)
        df = df[mask1]
        mask2 = (df["Created"] >= self.start_date) & (df["Created"] <= self.end_date)
        df = df[mask2]

        L1 = df.values.tolist()
        L1.insert(0, L[0])
        L = L1

        L = self.fetchUtils.add_final_assignment_group(L)
        L = self.fetchUtils.add_final_environment(L)
        L = self.fetchUtils.add_final_state(L)
        L = self.fetchUtils.fix_category_subcategory_blanks(L)

        # Write data to excel
        wb_out = Workbook()
        ws_out = wb_out.create_sheet("Incident", 0)
        for l in L:
            ws_out.append(l)
        try:
            wb_out.save(self.intermediary_incident_sheet_path)
        except Exception as e:
            print("File couldn't be written", e)

        return L

    def fetch_response_resolution_sheet_data(self):

        wb = openpyxl.load_workbook(self.response_resolution_sheet_path, data_only=True)
        ws = wb.active
        L = []
        for row in ws.rows:
            l = []
            for cell in row:
                l.append(cell.value)
            L.append(l)

        df = pd.DataFrame(L[1:], columns=L[0])
        mask1 = (df["Created"] >= self.start_date) & (df["Created"] <= self.end_date)
        mask2 = (df['Resolved'] >= self.start_date) & (df['Resolved'] <= self.end_date)
        df = df[mask1 | mask2]
        mask3 = (df["Project Key"] == self.project_key)
        df = df[mask3]

        L1 = df.values.tolist()
        L1.insert(0, L[0])
        L = L1

        L = self.fetchUtils.add_final_assignment_group(L)
        L = self.fetchUtils.add_final_environment(L)
        L = self.fetchUtils.add_final_state(L)
        L = self.fetchUtils.fix_category_subcategory_blanks(L)
        L = self.fetchUtils.add_bep(L)

        # Write data to excel
        wb_out = Workbook()
        ws_out = wb_out.create_sheet("Response Resolution", 0)
        for l in L:
            ws_out.append(l)
        try:
            wb_out.save(self.intermediary_response_resolution_sheet_path)
        except Exception as e:
            print("File couldn't be written", e)

        return L

    def fetch_change_request_sheet_data(self):
        wb = openpyxl.load_workbook(self.change_request_sheet_path, data_only=True)
        ws = wb.active
        L = []
        for row in ws.rows:
            l = []
            for cell in row:
                l.append(cell.value)
            L.append(l)

        # Apply filter for date range to ensure data is filtered accordingly
        df = pd.DataFrame(L[1:], columns=L[0])
        mask1 = (df["Created"] >= self.start_date) & (df["Created"] <= self.end_date)
        df = df[mask1]
        mask2 = (df["Project Key"] == self.project_key)
        df = df[mask2]

        L1 = df.values.tolist()
        L1.insert(0, L[0])
        L = L1

        L = self.fetchUtils.add_final_assignment_group(L)
        L = self.fetchUtils.fix_category_subcategory_blanks(L)

        # Writing to file
        wb_out = Workbook()
        ws_out = wb_out.create_sheet("Change Request", 0)
        for l in L:
            ws_out.append(l)
        try:
            wb_out.save(self.intermediary_change_request_sheet_path)
        except Exception as e:
            print("File couldn't be written.", e)

        return L

    def fetch_request_sheet_data(self):
        wb = openpyxl.load_workbook(self.request_sheet_path, data_only=True)
        ws = wb.active
        L = []
        for row in ws.rows:
            l = []
            for cell in row:
                l.append(cell.value)
            L.append(l)

        # Apply filter for date range to ensure data is filtered accordingly
        df = pd.DataFrame(L[1:], columns=L[0])
        # mask1 = (df["Created"] >= self.start_date) & (df["Created"] <= self.end_date)
        # df = df[mask1]
        mask2 = (df["Project Key"] == self.project_key)
        df = df[mask2]

        L1 = df.values.tolist()
        L1.insert(0, L[0])
        L = L1

        # Writing to file
        wb_out = Workbook()
        ws_out = wb_out.create_sheet("Request Items", 0)
        for l in L:
            ws_out.append(l)
        try:
            wb_out.save(self.intermediary_request_sheet_path)
        except Exception as e:
            print("File couldn't be written.", e)

        return L


"""
# Testing purpose only
fd = FetchData()
"""